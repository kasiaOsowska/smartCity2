from dataclasses import dataclass

from openpyxl.reader.excel import load_workbook


# Define variable to read sheet
class ExcelReader:
    data_file_CO = '2021_CO_1g.xlsx'
    data_file_NO2 = '2021_NO2_1g.xlsx'
    data_file_PM10 = '2021_PM10_1g.xlsx'

    start_row = 5

    # Gdynia
    # PorebskCO, PorebskNO2, PorebskPM10
    # SzafranCO - does not exist, no such data, SzafranNO2, SzafranPM10
    # Sopot
    # BiPlowcCO, BiPlowcNO2 - does not exist, no such data, BiPlowcPM10 - does not exist, no such data
    # Gdansk
    # PowWiel provides only data about PM2.5
    # WyzwoleCO, WyzwoleNO2, WyzwolePM10
    # LeczkowCO, LeczkowNO2, LeczkowPM10
    # PowWarsCO, PowWarsNO2, PowWarsPM10

    # CO
    # Gdynia
    PorebskCO = []
    # Sopot
    BiPlowcCO = []
    # Gdansk
    WyzwoleCO = []
    LeczkowCO = []
    PowWarsCO = []

    # NO2
    # Gdynia
    PorebskNO2 = []
    SzafranNO2 = []

    # Gdansk
    WyzwoleNO2 = []
    LeczkowNO2 = []
    PowWarsNO2 = []

    # PM10
    # Gdynia
    PorebskPM10 = []
    SzafranPM10 = []
    # Gdansk
    WyzwolePM10 = []
    LeczkowPM10 = []
    PowWarsPM10 = []
    
    def __init__(self):
        # All data is required to make statistics, so  it's better to load them once
        print("starting to load data...")

        wb = load_workbook(self.data_file_CO)
        name = wb.sheetnames[0]
        ws = wb[name]
        all_rows = list(ws.rows)

        index = 0
        # loading CO
        for row in all_rows:
            if index < 5:
                index += 1
                continue

            # stations: Szafran and PowWiel does not collect such data
            self.PorebskCO.insert(index, ExcelData(str(row[0].value), row[41].value))
            self.BiPlowcCO.insert(index, ExcelData(str(row[0].value), row[44].value))
            self.WyzwoleCO.insert(index, ExcelData(str(row[0].value), row[40].value))
            self.LeczkowCO.insert(index, ExcelData(str(row[0].value), row[38].value))
            self.PowWarsCO.insert(index, ExcelData(str(row[0].value), row[39].value))
            index = index + 1

        index = 0
        wb = load_workbook(self.data_file_NO2)
        name = wb.sheetnames[0]
        ws = wb[name]
        all_rows = list(ws.rows)

        # loading NO2
        for row in all_rows:
            if index < 5:
                index += 1
                continue

            # stations: PowWiel, BiPlowc does not collect such data
            self.SzafranNO2.insert(index, ExcelData(str(row[0].value), row[93].value))
            self.PorebskNO2.insert(index, ExcelData(str(row[0].value), row[92].value))
            self.WyzwoleNO2.insert(index, ExcelData(str(row[0].value), row[91].value))
            self.PowWarsNO2.insert(index, ExcelData(str(row[0].value), row[90].value))
            self.LeczkowNO2.insert(index, ExcelData(str(row[0].value), row[89].value))

            index = index + 1

        index = 0
        wb = load_workbook(self.data_file_NO2)
        name = wb.sheetnames[0]
        ws = wb[name]
        all_rows = list(ws.rows)

        # loading PM10
        for row in all_rows:
            if index < 5:
                index += 1
                continue

            # stations: PowWiel, BiPlowc does not collect such data
            self.SzafranPM10.insert(index, ExcelData(str(row[0].value), row[118].value))
            self.PorebskPM10.insert(index, ExcelData(str(row[0].value), row[117].value))
            self.WyzwolePM10.insert(index, ExcelData(str(row[0].value), row[116].value))
            self.PowWarsPM10.insert(index, ExcelData(str(row[0].value), row[115].value))
            self.LeczkowPM10.insert(index, ExcelData(str(row[0].value), row[114].value))

            index = index + 1


@dataclass
class ExcelData:
    date: str
    amount: float

    def __init__(self, date, amount):
        self.date = date
        self.amount = amount
